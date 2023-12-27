from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from rich import print, pretty
import time
import requests
import urllib.request
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from  bs4 import BeautifulSoup
import xlrd
from openpyxl import load_workbook
from itertools import cycle

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from google.oauth2 import service_account

GENERARRUT = False
BASEURL = 'https://elrutificador.com/'

from selenium.webdriver.chrome.options import Options

options = Options()
options.add_experimental_option("detach", True)


def calcular_digito_verificador(rut):
    reversed_digits = map(int, reversed(str(rut)))
    factors = cycle(range(2, 8))
    s = sum(d * f for d, f in zip(reversed_digits, factors))
    return 'K' if s % 11 == 1 else str((11 - s % 11) % 10)

def generar_y_escribir_ruts(rango_inicio, rango_fin, nombre_archivo, tamaño_lote=10000):
    with open(nombre_archivo, 'w') as archivo:
        for rut in range(rango_inicio, rango_fin + 1):
            archivo.write(f"{rut}-{calcular_digito_verificador(rut)}\n")
            if rut % tamaño_lote == 0:
                print(f"Procesados {rut} RUTs")

def validar_rut(rut):
    
    rut = rut.replace(".", "").replace("-", "")
    

    cuerpo, dv = rut[:-1], rut[-1].upper()

    
    if not cuerpo.isdigit() or dv not in "0123456789K":
        return False
    

    suma = sum(int(digit) * factor for digit, factor in zip(reversed(cuerpo), cycle(range(2, 8))))
    restante = 11 - suma % 11
    dv_calculado = {10: 'K', 11: '0'}.get(restante, str(restante))
    

    return dv == dv_calculado


rango_inicio = 26000000  
rango_fin = 28000000 
nombre_archivo = "ruts_generados.txt"




SHEET = 'sheet-rutificador@rutificador-409414.iam.gserviceaccount.com'
KEY_SHEET = 'rutificador-409414-45b7daca2bd2.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
KEY = '45b7daca2bd24ad6ec37c24815fc635a85d21744'



if GENERARRUT:
    generar_y_escribir_ruts(rango_inicio, rango_fin, nombre_archivo)




def checkElement(el):
    el = len(el)
    if el > 0:
        return 1
    else:
        return 0




driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get(BASEURL)
workbook = load_workbook("rut.xlsx")
sheet = workbook.active
ruts = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]

print(f"Se encontraron {len(ruts)} RUTs en el archivo")



workbook_openpyxl = load_workbook("rut.xlsx")
sheet_openpyxl = workbook_openpyxl.active
datos_ruts = []
fechas_nacimiento = []


for i, rut in enumerate(ruts, start=1):
    time.sleep(2)

    if not validar_rut(rut):
        print(f"RUT {rut} no es válido.")
        continue
    try:    
        driver.find_element("xpath", '//*[@id="btrut"]').click()
        driver.find_element("xpath", '//*[@id="txtrut"]').send_keys(str(rut))
        driver.find_element("xpath", '//*[@id="btnsrut"]').click()
        fecha_nacimiento_element = driver.find_element(By.XPATH, "//td[b='F. Nacimiento']/following-sibling::td")
        fecha_nacimiento = fecha_nacimiento_element.text

        print(f"RUT {i}: {rut} - Fecha de nacimiento: {fecha_nacimiento}")
        time.sleep(2)
        
        driver.back()
        time.sleep(2)

    
    except NoSuchElementException:
        fecha_nacimiento = "No encontrada"
    
    except Exception as e:
        print(f"Se encontró un error al procesar el RUT {rut}: {e}")
        fecha_nacimiento = "Error"
    
    sheet.cell(row=i+1, column=2).value = fecha_nacimiento
        
workbook.save(filename="rut.xlsx")

print(f"Datos extraídos y guardados en ")



